import configparser
import os
import re
import sys

import openpyxl
import requests
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def get_cookie(need_new=False):
    conf = configparser.ConfigParser()
    conf.read('config.ini', encoding='utf-8')
    if not conf.has_section('Storage'):
        conf.add_section('Storage')
        print('首次运行请输入 cookie，建议从胡桃工具箱(https://hut.ao/)获取，详情查看原帖')
    if not conf.has_option('Storage', 'cookie') or need_new:
        print('cookie', end=': ')
        cookie = input().strip()
        print()
        conf.set("Storage", 'cookie', cookie)
        with open("config.ini", "w+") as f:
            conf.write(f)
    return conf.get("Storage", 'cookie')


def get_json(share_code, cookie):
    url = 'https://api-takumi.mihoyo.com/event/e20200928calculate/v1/furniture/blueprint'
    params = {
        'share_code': share_code,
        'region': 'cn_gf01'
    }
    headers = {
        'cookie': cookie
    }
    try:
        return requests.get(url, params, headers=headers).json()
    except requests.exceptions.SSLError:
        exit_by_tips('请关闭系统代理后重试')


def parse_json(data):
    # cookie无效
    if data['retcode'] == -100:
        print('cookie无效，请重新获取')
        get_cookie(True)
        get_furniture_list()
    # 其他错误
    elif not data['retcode'] == 0:
        exit_by_tips(data['message'])
    res = []
    for furniture in data['data']['list'] + data['data']['not_calc_list']:
        jump = ''
        if furniture["wiki_url"] != "":
            url = re.sub(r'\?.*$', '', furniture["wiki_url"])
            jump = '=HYPERLINK(\"' + url + '\", \"跳转\")'
        res.append([furniture['id'], furniture['name'], "", furniture['num'], furniture['level'], jump])
        # print(furniture['name'], furniture['num'], url, sep='\t')
    res.sort(key=lambda cur_res: (cur_res[4], cur_res[3], cur_res[0]))
    return res


def out_excel(data, share_code):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['摆设ID', '摆设名称', '拥有数量', '所需数量', '星级', 'Wiki'])
    for row in data:
        ws.append(row)

    # 稀有度和拥有数量达标的条件格式
    rule = ColorScaleRule(start_type='num', start_value=2, start_color='ffc9e8a8',
                          mid_type='num', mid_value=3, mid_color='ff6ddbff',
                          end_type='num', end_value=4, end_color='ffcaa8e5')
    ws.conditional_formatting.add(f'E2:E{len(data) + 1}', rule)
    green_fill = PatternFill(start_color='e2efda',
                             end_color='e2efda',
                             fill_type='solid')
    ws.conditional_formatting.add(f'C2:C{len(data) + 1}',
                                  CellIsRule(operator='greaterThanOrEqual', formula=['$D2'], fill=green_fill))

    # 所需数量数据条
    rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=max(row[3] for row in data),
                       color="FFfcd5b4")
    ws.conditional_formatting.add(f'D2:D{len(data) + 1}', rule)

    # 边框、字体和居中
    side = Side(style='thin', color='00000000')
    border = Border(*(4 * [side]))
    font = Font(name='等线')
    alignment = Alignment(horizontal='center', vertical='center')
    rows = ws[f'A1:F{len(data) + 1}']
    for row in rows:
        for cell in row:
            cell.border = border
            cell.font = font
            cell.alignment = alignment

    # 摆设名称列宽
    ws.column_dimensions['B'].width = 25

    # 冻结首行
    ws.freeze_panes = 'A2'

    try:
        wb.save(f'{share_code}.xlsx')
    except PermissionError:
        exit_by_tips('请关闭当前Excel表后重试')


def exit_by_tips(s):
    print()
    print(s)
    print('按 Enter 键退出程序')
    input()
    sys.exit()


def get_furniture_list():
    cookie = get_cookie()
    print('洞天摹数', end=': ')
    share_code = input().strip()
    json_data = get_json(share_code, cookie)
    res = parse_json(json_data)
    out_excel(res, share_code)
    print(f'获取完成，正在打开 {share_code}.xlsx\n')
    os.startfile(f'{share_code}.xlsx')


if __name__ == '__main__':
    print('项目地址: https://github.com/lingkai5wu/getYsFurnitureList')
    print('Tips: 可右键直接粘贴最近复制的内容\n')
    get_furniture_list()
